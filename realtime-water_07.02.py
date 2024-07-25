from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime
import os
from openpyxl import Workbook

def get_water_l():
    driver.get("https://sqfb.slt.zj.gov.cn/weIndex.html#/main/map/realtime-water/")
    input()
    for j in range(1,31):
        print('start')
        for i in range(1,7):
            name = str(j)+'_pic_' + str(i)+'.jpg'
            driver.get_screenshot_as_file(os.path.join(r'E:\zhejiang_university\srtp\project01\word\pic', name))
            print(name+' is ok')
            time.sleep(6)

def read_txt(loc):
    with open(loc, 'r', encoding='utf-8') as file:
        data_list = file.readlines()
    des = [line.strip() for line in data_list]
    return des

def get_water(g_type):
    water_data = []
    hour_data = []
    rain_data = []
    driver.get("https://sqfb.slt.zj.gov.cn/weIndex.html#/main/map/realtime-water/")
    time.sleep(7)
    #确定区域
    # elem = driver.find_element_by_xpath('//ul[@class="el-scrollbar__view el-select-dropdown__list"]')
    # elem.find_element_by_xpath('/html/body/div[6]/div[1]/div[1]/ul/li[3]')
    # driver.execute_script("arguments[0].scrollIntoView(true);", elem)
    if g_type == 1:
        i = 2
        j = 19
    elif g_type == 2:
        i = 11
        j = 23
    # 确定市
    elems = driver.find_elements_by_xpath('//li[@class="el-select-dropdown__item"]')
    driver.execute_script("arguments[0].click();", elems[i])
    time.sleep(1.5)
    # 确定县
    elems = driver.find_elements_by_xpath('//li[@class="el-select-dropdown__item"]')
    driver.execute_script("arguments[0].click();", elems[j])
    time.sleep(1.5)

    # 选择类型
    elems = driver.find_elements_by_xpath('//label[@class="el-checkbox custom-checkbox"]')
    for elem in elems:
        driver.execute_script("arguments[0].click();", elem)
        time.sleep(0.5)
    time.sleep(2)

    # 选择“所有”
    elem = driver.find_element_by_xpath("/html/body/div[3]/div/div/div[1]/div[3]/div[1]/div[2]/div/div[4]/div[1]/span[1]")
    driver.execute_script("arguments[0].click();", elem)
    time.sleep(1.5)

    count = count_strings(des)

    n_flag = 0
    for d in des:
        if d != des[0]:
            driver.switch_to.parent_frame()
        # 输入站点
        elem = driver.find_element_by_xpath("/html/body/div[3]/div/div/div[1]/div[3]/div[1]/div[2]/div/div[4]/div[2]/input")
        elem.clear()
        # 键入关键词
        elem.send_keys(d)
        elem.send_keys(Keys.ENTER)
        time.sleep(1.5)

        # 点击搜索
        elem = driver.find_element_by_xpath('//button[@type="button"]')
        driver.execute_script("arguments[0].click();", elem)
        time.sleep(1.5)
        water = []
        hour = []
        rain = []
        print('\n' + d)
        #进入站点详情
        try:
            driver.switch_to.frame('myframe')
            elems = driver.find_elements_by_xpath('//div[@class="cell text-c"]')
            flag = 0
            while True:
                # 有的完全匹配站点会被放在下面
                if elems[2 + 5 * (count[d]+flag)].text != d:
                    flag+=1
                else:
                    driver.execute_script("arguments[0].click();", elems[2+5*(count[d]+flag)])
                    break
            time.sleep(5)

            #切换天数
            driver.switch_to.parent_frame()
            driver.switch_to.frame('myMarkframe')
            elems = driver.find_elements_by_xpath('//li[@class="el-select-dropdown__item"]')
            driver.execute_script("arguments[0].click();", elems[1])
            time.sleep(5)

            # 进入表格
            driver.find_elements_by_xpath('//tr[@class="el-table__row"]')
            for step in range(1,73):
                xpath = '//*[@id="pane-ssgc"]/div[2]/div/div[4]/div[2]/table/tbody/tr'
                xpath = xpath+'['+str(step)+']/td[2]/div'
                hour_elem = driver.find_elements_by_xpath(xpath)

                xpath3 = '//*[@id="pane-ssgc"]/div[2]/div/div[3]/table/tbody/tr'
                xpath3 = xpath3+'['+str(step)+']/td[3]/div'
                rain_elem = driver.find_elements_by_xpath(xpath3)

                css = '#pane-ssgc > div.listTb > div > div.el-table__body-wrapper.is-scrolling-none > table > tbody > tr:nth-child'
                css += "("+str(step)+') > td.el-table_1_column_4.el-table__cell > div'
                driver.find_elements_by_css_selector(css)
                xpath2 = '//*[@id="pane-ssgc"]/div[2]/div/div[3]/table/tbody/tr'
                xpath2 += '['+str(step)+']/td[4]/div/span'
                water_elem = driver.find_elements_by_xpath(xpath2)

                water.append(water_elem[0].text)
                hour.append(hour_elem[0].text)
                rain.append(rain_elem[0].text)
            print(hour)
            print(rain)
            print(water)
        except:
            print('None!')
            n_flag += 1
            pass
        water_data.append(water)
        hour_data.append(hour)
        rain_data.append(rain)
        count[d] += 1
        time.sleep(3)
    if n_flag:
        print('something wrong')
    save_exl(des, water_data, hour_data, 'water',g_type)
    save_exl(des, rain_data, hour_data, 'rain',g_type)


def save_exl(title,data_list,hour_list,name,type):
    path = r'E:\zhejiang_university_new\code\rain-water'
    path = os.path.join(path, str(type)+'_'+name+'_'+hour_list[0][0][0:8]+'.xlsx')
    wb = Workbook()
    ws = wb.active
    for j in range(len(hour_list[0])):
        ws.cell(3 + j, 1, '2024-'+hour_list[0][len(hour_list[0])-1-j])
    for i in range(len(data_list)):
        ws.cell(1, i + 2, title[i])
        for j in range(len(data_list[i])):
            ws.cell(j+3, i+2, data_list[i][len(data_list[i])-1-j])
    wb.save(path)

def count_strings(lst):
    count_dict = {}
    for s in lst:
        count_dict[s] = 0
    return count_dict

if __name__ == '__main__':
    driver = webdriver.Chrome \
        ('C:/Program Files/Google/Chrome/Application/chromedriver')
    driver.set_window_size(1000, 1550)
    driver.set_window_position(0, 0)
    # 1:杭州钱塘 2：台州临海大田平原
    for i in range(2):
        print('\n')
        print(datetime.now())
        j = i+1
        if j == 1:
            path = r'E:\zhejiang_university_new\code\rain-water\1_qiantang\location.txt'
        else:
            path = r'E:\zhejiang_university_new\code\rain-water\2_datian\location.txt'
        des = read_txt(path)
        get_water(j)
